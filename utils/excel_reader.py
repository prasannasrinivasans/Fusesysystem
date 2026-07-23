import csv
import os
from pathlib import Path
from faker import Faker
from openpyxl import Workbook, load_workbook

def get_excel_data(file_path, sheet_name=None):

    data = []

    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(tuple(str(cell) if cell is not None else "" for cell in row))

    return data


def save_excel_data(file_path, rows, sheet_name="Sheet1"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    headers = [
        "first_name",
        "last_name",
        "address2",
        "address3",
        "city",
        "state",
        "zipcode",
        "ssn",
        "account",
        "gender",
        "chcp"
    ]
    sheet.append(headers)

    for row in rows:
        sheet.append(row)

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(file_path)


def generate_fake_member_data(count=1, locale="en_US"):
    faker = Faker(locale)
    rows = []

    for _ in range(count):
        first_name = faker.first_name()
        last_name = faker.last_name()
        address2 = faker.street_address()
        address3 = faker.secondary_address()
        city = faker.city()
        state = faker.state_abbr()
        zipcode = faker.zipcode()
        ssn = faker.ssn()
        account = faker.numerify(text="##########")
        gender = faker.simple_profile()["sex"]
        chcp = faker.random_element(elements=("Y", "N"))

        rows.append((
            first_name,
            last_name,
            address2,
            address3,
            city,
            state,
            zipcode,
            ssn,
            account,
            gender,
            chcp
        ))

    return rows